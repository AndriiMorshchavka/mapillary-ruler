# This program makes Mapillary request by Mapillary id and saves data to geojson as well as downloads and saves the original image

import requests
import os
import openpyxl

# Specify output directory
directory = os.path.abspath(__file__)
excel_path = os.path.join (directory, r"..//1_new_images.xlsx")
images_path = os.path.join (directory, r"..//new_images")

# Define Mapillary API base URL
base_url = "https://graph.mapillary.com"

# Specify request parameters
access_token = "MLY|7001510073286829|ea90511ea191e05c31eddbbf6247bf93"
fields = "atomic_scale,thumb_original_url,computed_geometry,computed_compass_angle,computed_rotation,camera_parameters,camera_type,height,width,exif_orientation,computed_altitude,creator,make,model"
print ("Type the pKey")
pKey = input()
print (f"You have typed pKey: {pKey}")

# Make Mapillary API request and pprocess it
response = requests.get(f"{base_url}/{pKey}?access_token={access_token}&fields={fields}")
if response.status_code == 200:
    print ("Mapillary request succesful")
    # Parse and process the response JSON
    data = response.json()
    # Access the thumb_original_url directly from the data dictionary
    image_url = data.get('thumb_original_url')
    camera_type = data.get ("camera_type")
    exif_orientation = data.get ("exif_orientation")
    if camera_type!="perspective" or exif_orientation != 1:
        print ("Image does not meet criteria")
    else:
        # Make request to download image
        response_image = requests.get(image_url)
        if response_image.status_code == 200:
            image_name = f"{pKey}.jpg"
            image_unique_path = os.path.join(images_path, image_name)
            with open(image_unique_path, 'wb') as f:
                f.write(response_image.content)
            print(f"Image successfully downloaded: {image_unique_path}")

            thumb_original_url = data.get('thumb_original_url')
            
            # Computed geometry coordinates
            computed_geometry = data.get("computed_geometry", {})
            computed_coordinates = computed_geometry.get("coordinates", [None, None])
            computed_lat = computed_coordinates[1]
            computed_lon = computed_coordinates[0]
            
            # Retrieve compass angle
            compass_angle = data.get("computed_compass_angle")

            # Retrieve altitude [from sea level?]
            altitude = data.get ("computed_altitude")

            # Retrieve rotation data
            computed_rotation = data.get("computed_rotation", [None, None, None])
            rotation_x = computed_rotation[0]
            rotation_y = computed_rotation[1]
            rotation_z = computed_rotation[2]

            # Camera parameters include focal length and distortion
            camera_parameters = data.get("camera_parameters", [None, None, None])
            focal_length = camera_parameters[0]
            k1 = camera_parameters[1]
            k2 = camera_parameters[2]

            # Retrieve image size
            height = data.get("height")
            width = data.get("width")
            
            # Retrieve atomic scale
            atomic_scale = data.get ("atomic_scale")
            
            # Retrieve username of creator, make and model of their camera
            creator = data.get ("creator")
            username = creator.get ("username")
            make = data.get ("make")
            model = data.get ("model")

            # Open the Excel file
            wb = openpyxl.load_workbook(excel_path)

            # Select the appropriate worksheet (assuming it's the first one)
            ws = wb.active

            # Find the next free row in the first column
            next_row = 2
            while ws.cell(row=next_row, column=2).value is not None:
                next_row += 1

            column_counter = 2
            # Write the image ID into the next free row in the first column
            ws.cell(row=next_row, column=column_counter).value = pKey
            column_counter += 1
            # Write the data into the table
            ws.cell(row=next_row, column=column_counter).value = thumb_original_url
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = computed_lat
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = computed_lon
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = altitude
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = compass_angle
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = camera_type
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = rotation_x
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = rotation_y
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = rotation_z
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = focal_length
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = k1
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = k2
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = height
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = width
            column_counter += 1
            ws.cell(row=next_row, column=column_counter).value = exif_orientation
            column_counter += 1
            ws.cell (row = next_row, column = column_counter).value = atomic_scale
            column_counter += 1
            ws.cell (row = next_row, column = column_counter).value = username
            column_counter += 1
            ws.cell (row = next_row, column = column_counter).value = make
            column_counter += 1
            ws.cell (row = next_row, column = column_counter).value = model
            column_counter += 1
            print (f"Successfully written data for the image with pKey {pKey} into Excel file")
            # Save the changes to the Excel file
            wb.save(excel_path)
        else:
            print(f"Failed to retrieve image from {image_url}")
else:
    print (f'Request failed with status code {response.status_code}: {response.text}')    